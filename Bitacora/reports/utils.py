import matplotlib.pyplot as plt
import io
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image, Table, TableStyle
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from ..models import Produccion, Paro, SlitterProduccion, SlitterParo, OsiladoraProduccion, OsiladoraParo

# Definir los tiempos disponibles en minutos para cada línea y turno
TIEMPOS_DISPONIBLES = {
    'Produccion': {'Matutino': 685, 'Vespertino': 420, 'Nocturno': 480},
    'Slitter': {'Matutino': 685, 'Vespertino': 420, 'Nocturno': 480},
    'Osciladora': {'Matutino': 700, 'Nocturno': 480},
    'Komatsu': {'Matutino': 685, 'Vespertino': 420, 'Nocturno': 480},
    'Seyi': {'Matutino': 685, 'Vespertino': 420, 'Nocturno': 485},
    'TND 200': {'Matutino': 535, 'Nocturno': 475},
    'Tanden 300': {'Matutino': 535, 'Vespertino': 425, 'Nocturno': 475},
}

# Función para calcular el GSPH (Golpes por Hora) para una línea de producción genérica
def calcular_gsph(producciones, linea):
    for produccion in producciones:
        tiempo_disponible = TIEMPOS_DISPONIBLES[linea].get(produccion.turno, 0)
        tiempo_total_horas = tiempo_disponible / 60  # Convertir minutos disponibles a horas
        if tiempo_total_horas > 0:
            produccion.gsph = produccion.piezas_ok / tiempo_total_horas
        else:
            produccion.gsph = 0
        produccion.save()

# Función para calcular el GSPH para la línea de producción "Slitter"
def calcular_gsph_slitter(slitter_producciones):
    for produccion in slitter_producciones:
        tiempo_disponible = TIEMPOS_DISPONIBLES['Slitter'].get(produccion.turno, 0)
        tiempo_total_horas = tiempo_disponible / 60  # Convertir minutos disponibles a horas
        if tiempo_total_horas > 0:
            produccion.gsph = produccion.kilos_ok / tiempo_total_horas
        else:
            produccion.gsph = 0
        produccion.save()

# Función para calcular el GSPH para la línea de producción "Osciladora"
def calcular_gsph_osiladora(osiladora_producciones):
    for produccion in osiladora_producciones:
        tiempo_disponible = TIEMPOS_DISPONIBLES['Osciladora'].get(produccion.turno, 0)
        tiempo_total_horas = tiempo_disponible / 60  # Convertir minutos disponibles a horas
        if tiempo_total_horas > 0:
            produccion.gsph = produccion.kilos_ok / tiempo_total_horas
        else:
            produccion.gsph = 0
        produccion.save()

# Función para generar el reporte PDF por turno
def generar_reporte_por_turno(response, turnos, linea):
    doc = SimpleDocTemplate(response, pagesize=letter)  # Crear el documento PDF
    styles = getSampleStyleSheet()  # Obtener los estilos para el documento
    elements = []  # Lista para almacenar los elementos del PDF

    # Agregar el título del reporte
    elements.append(Paragraph('Reporte de Bitácoras', styles['Title']))
    elements.append(Spacer(1, 12))  # Espaciador para separar los elementos

    # Iterar sobre cada turno para generar los datos del reporte
    for turno in turnos:
        if linea == 'Slitter':
            producciones = SlitterProduccion.objects.filter(turno=turno)
            calcular_gsph_slitter(producciones)
        elif linea == 'Osciladora':
            producciones = OsiladoraProduccion.objects.filter(turno=turno)
            calcular_gsph_osiladora(producciones)
        else:
            producciones = Produccion.objects.filter(turno=turno, linea=linea)
            calcular_gsph(producciones, linea)

        paros = Paro.objects.filter(turno=turno, linea=linea)

        if producciones.exists() or paros.exists():
            elements.append(Paragraph(f'Turno: {turno}', styles['Heading2']))
            elements.append(Spacer(1, 12))

            # Definir la estructura de la tabla de datos
            data = [['Línea', 'Tiempo Disponible (min)', 'GSPH', 'Objetivo', 'Top 3 Tiempos de Inactividad', 'Inactividad (min)', '% Inactividad Disponible', 'Producción (Kg/pcs)', 'No. de Rollos Procesados', 'No. de Cambios de Modelo', 'Tiempo de Cambio de Modelo']]

            for produccion in producciones:
                top_inactividad = "\n".join([paro.descripcion for paro in paros.order_by('-duracion')[:3]])
                data.append([
                    produccion.modelo.nombre if hasattr(produccion, 'modelo') else produccion.linea,
                    TIEMPOS_DISPONIBLES[linea].get(turno, 0),
                    produccion.gsph,
                    produccion.objetivo,
                    top_inactividad,
                    sum([paro.duracion for paro in paros]),
                    produccion.porcentaje_inactividad_disponible,
                    produccion.piezas_ok if hasattr(produccion, 'piezas_ok') else produccion.kilos_ok,
                    '-',
                    '-',
                    '-'
                ])

            # Crear la tabla con los datos recopilados
            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),  # Fondo de la cabecera en gris
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),  # Texto de la cabecera en blanco
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  # Alineación centrada para todas las celdas
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),  # Fuente de la cabecera en negrita
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),  # Relleno inferior de la cabecera
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),  # Fondo de las celdas en beige
                ('GRID', (0, 0), (-1, -1), 1, colors.black),  # Rejilla negra para todas las celdas
            ]))
            elements.append(table)
            elements.append(Spacer(1, 24))

            # Generar gráfico para el turno
            departamentos = [produccion.departamento for produccion in producciones]
            inactividad = [sum([paro.duracion for paro in paros])]

            plt.figure(figsize=(10, 5))
            plt.bar(departamentos, inactividad, color='blue')
            plt.xlabel('Departamento')
            plt.ylabel('Inactividad (min)')
            plt.title(f'Inactividad por Departamento - Turno {turno}')

            buffer = io.BytesIO()
            plt.savefig(buffer, format='png')
            buffer.seek(0)

            elements.append(Image(buffer, width=500, height=250))
            elements.append(Spacer(1, 48))

    # Generar el documento PDF con los elementos recopilados
    doc.build(elements)

# Función para generar el reporte Excel por turno
def generar_reporte_excel(turnos, linea):
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Bitácoras"

    # Encabezados
    headers = ['Línea', 'Tiempo Disponible (min)', 'GSPH', 'Objetivo', 'Top 3 Tiempos de Inactividad', 'Inactividad (min)', '% Inactividad Disponible', 'Producción (Kg/pcs)', 'No. de Rollos Procesados', 'No. de Cambios de Modelo', 'Tiempo de Cambio de Modelo']
    ws.append(headers)

    # Aplicar estilo a los encabezados
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[col_letter + '1'].font = Font(bold=True)

    for turno in turnos:
        if linea == 'Slitter':
            producciones = SlitterProduccion.objects.filter(turno=turno)
            calcular_gsph_slitter(producciones)
        elif linea == 'Osciladora':
            producciones = OsiladoraProduccion.objects.filter(turno=turno)
            calcular_gsph_osiladora(producciones)
        else:
            producciones = Produccion.objects.filter(turno=turno, linea=linea)
            calcular_gsph(producciones, linea)

        paros = Paro.objects.filter(turno=turno, linea=linea)

        for produccion in producciones:
            top_inactividad = ", ".join([paro.descripcion for paro in paros.order_by('-duracion')[:3]])
            ws.append([
                produccion.modelo.nombre if hasattr(produccion, 'modelo') else produccion.linea,
                TIEMPOS_DISPONIBLES[linea].get(turno, 0),
                produccion.gsph,
                produccion.objetivo,
                top_inactividad,
                sum([paro.duracion for paro in paros]),
                produccion.porcentaje_inactividad_disponible,
                produccion.piezas_ok if hasattr(produccion, 'piezas_ok') else produccion.kilos_ok,
                '',
                '',
                ''
            ])

    wb.save('reporte_bitacoras.xlsx')
